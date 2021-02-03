import xlsxwriter
import openpyxl

import numpy as np

from perceptron import Perceptron

def save_model(model, file_name):
	""" save model weights and biases in excel format """
	workbook = xlsxwriter.Workbook(file_name)
	for i in range(model.nLayers-1):
		worksheet = workbook.add_worksheet("Weights "+str(i))
		for j in range(len(model.weights[i])):
			for k in range(len(model.weights[i][j])):
				worksheet.write(j, k, model.weights[i][j][k])
		worksheet = workbook.add_worksheet("Biases "+str(i))
		for j in range(len(model.biases[i])):
			worksheet.write(j, 0, model.biases[i][j])
	workbook.close()
	
def load_xlsx(file_name, layout):
	""" import model weights and biases from excel file """
	model = Perceptron(layout)
	wb = openpyxl.load_workbook(file_name)
	sheets = wb.sheetnames
	weights = []
	biases = []		
	for i in range(model.nLayers-1):
		sheet = wb[sheets[i*2+1]]
		temp_bias = []
		for j in range(layout[i+1]):
			temp_bias.append(np.float64(sheet.cell(row=j+1, column=1).value))
		biases.append(np.array(temp_bias))
		sheet = wb[sheets[i*2]]
		temp_weight = np.empty((layout[i+1], layout[i]))
		for j in range(layout[i+1]):
			for k in range(layout[i]):
				temp_weight[j, k] = np.float64(sheet.cell(row=j+1, column=k+1).value)
		weights.append(np.array(temp_weight))
	model.weights=weights
	model.biases=biases
	return model	